import pandas as pd
import numpy as np
import scanpy as sc
import openpyxl
import matplotlib.pyplot as plt
from sklearn.metrics import roc_auc_score

# ── 1. Load count matrix ──────────────────────────────────────────────────────
print("Loading data...")
df = pd.read_csv(
    "all_cells_counts.txt",
    sep="\t",
    comment="#",
    index_col=0
)
df = df.drop(columns=["Chr", "Start", "End", "Strand", "Length", "gene_id"])
df.columns = [c.split("/")[-1].replace("_sorted.bam", "") for c in df.columns]

# ── 2. Load metadata and filter to organoid cells ─────────────────────────────
meta = pd.read_csv("SraRunTable.csv", index_col="Run")
organoid_mask = meta["tissue"].isin([
    "Dissociated whole cerebral organoid",
    "Microdissected cortical-like ventricle from cerebral organoid"
])
organoid_srr = [s for s in meta[organoid_mask].index if s in df.columns]
df = df[organoid_srr].T  # cells x genes

# ── 3. Log normalise ──────────────────────────────────────────────────────────
df_norm = df.div(df.sum(axis=1), axis=0) * 1e4
df_log = np.log1p(df_norm)

# ── 4. Load PCA genes from Dataset S2 ────────────────────────────────────────
wb = openpyxl.load_workbook("pnas.1520760112.sd02.xlsx")
ws = wb['PCA_genes']
pca_genes = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
pca_genes_present = [g for g in pca_genes if g in df_log.columns]
print(f"PCA genes found in our data: {len(pca_genes_present)}")

# ── 5. Load cluster assignments from scanpy results ───────────────────────────
# We need to rerun the clustering to get leiden assignments
# Build AnnData and recluster
print("Rebuilding clusters...")
import scanpy as sc

adata = sc.AnnData(df_log)

# Filter
sc.pp.filter_cells(adata, min_genes=200)
sc.pp.filter_genes(adata, min_cells=3)

# Filter housekeeping genes
actb = adata[:, "ACTB"].X.flatten() if "ACTB" in adata.var_names else np.ones(adata.n_obs)
gapdh = adata[:, "GAPDH"].X.flatten() if "GAPDH" in adata.var_names else np.ones(adata.n_obs)
adata = adata[(actb > 0) & (gapdh > 0)].copy()

# Subset to PCA genes
pca_genes_filtered = [g for g in pca_genes_present if g in adata.var_names]
adata_pca = adata[:, pca_genes_filtered].copy()

sc.pp.scale(adata_pca, max_value=10)
sc.tl.pca(adata_pca, svd_solver="arpack")
sc.pp.neighbors(adata_pca, n_neighbors=10, n_pcs=30)
sc.tl.leiden(adata_pca, resolution=1.5, flavor='igraph',
             n_iterations=2, directed=False)

clusters = adata_pca.obs['leiden']
print(f"Clusters found: {clusters.nunique()}")

# Get expression matrix for PCA genes only
X = adata_pca.X  # scaled expression
genes = adata_pca.var_names.tolist()
cluster_ids = sorted(clusters.unique(), key=lambda x: int(x))

# ── 6. Run ROC test ───────────────────────────────────────────────────────────
print("\nRunning ROC test on PCA genes...")
results = []

for cluster in cluster_ids:
    y = (clusters == cluster).astype(int).values
    print(f"  Cluster {cluster} ({y.sum()} cells)...")
    
    for i, gene in enumerate(genes):
        expr = X[:, i]
        if hasattr(expr, 'toarray'):
            expr = expr.toarray().flatten()
        
        if expr.max() == expr.min():
            continue
            
        try:
            auc = roc_auc_score(y, expr)
            power = 2 * abs(auc - 0.5)
            results.append({
                'cluster': f'C{int(cluster)+1}',
                'gene': gene,
                'auc': round(auc, 3),
                'power': round(power, 3)
            })
        except Exception:
            continue

roc_df = pd.DataFrame(results)
print(f"\nTotal ROC results: {len(roc_df)}")

# ── 7. Get top 10 genes per cluster by power ──────────────────────────────────
top_genes = roc_df.groupby('cluster').apply(
    lambda x: x.nlargest(10, 'power')
).reset_index(drop=True)

print("\nTop 10 marker genes per cluster (by classification power):")
for cluster in [f'C{i+1}' for i in range(clusters.nunique())]:
    cg = top_genes[top_genes['cluster'] == cluster]
    genes_list = cg[['gene', 'auc', 'power']].values.tolist()
    print(f"\n{cluster}:")
    for g, auc, power in genes_list:
        print(f"  {g}: AUC={auc:.3f}, power={power:.3f}")

# ── 8. Compare to paper's Dataset S2 AUC scores ───────────────────────────────
print("\n\nComparing to paper's Dataset S2 marker genes...")

paper_cell_types = {
    'Neurons': 'neurons_all',
    'Dorsal NPC': 'dorsal_cortex_progenitors',
    'Dorsal neuron': 'dorsal_cortex_neurons',
    'Ventral NPC': 'ventral_progenitors',
    'Hem/RSPO': 'RSPO_cells',
    'Mesenchyme': 'mesenchyme_cells',
}

paper_markers = {}
for label, sheet in paper_cell_types.items():
    ws = wb[sheet]
    genes_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] and row[3]:
            genes_data.append({
                'gene': row[0],
                'paper_auc': row[1],
                'paper_power': row[3]
            })
    paper_markers[label] = pd.DataFrame(genes_data)

# For each cluster, find overlap with paper markers and compare AUC
print("\nCluster correspondence to paper cell types:")
print("="*60)

correspondence = {}
for cluster in [f'C{i+1}' for i in range(clusters.nunique())]:
    our_top = top_genes[top_genes['cluster'] == cluster][['gene', 'auc', 'power']]
    our_genes = set(our_top['gene'].tolist())
    
    best_match = None
    best_overlap = 0
    
    for cell_type, paper_df in paper_markers.items():
        paper_genes = set(paper_df['gene'].tolist())
        overlap = len(our_genes & paper_genes)
        if overlap > best_overlap:
            best_overlap = overlap
            best_match = cell_type
    
    correspondence[cluster] = {
        'best_match': best_match,
        'overlap': best_overlap
    }
    print(f"{cluster}: best match = {best_match} ({best_overlap} shared top genes)")

# ── 9. Save results ───────────────────────────────────────────────────────────
roc_df.to_csv("roc_results_all.csv", index=False)
top_genes.to_csv("roc_top_genes_per_cluster.csv", index=False)
print("\nSaved: roc_results_all.csv")
print("Saved: roc_top_genes_per_cluster.csv")

# ── 10. Plot heatmap of AUC overlap with paper cell types ─────────────────────
print("\nPlotting AUC overlap heatmap...")

cluster_labels = [f'C{i+1}' for i in range(clusters.nunique())]
ct_labels = list(paper_cell_types.keys())

overlap_matrix = pd.DataFrame(0.0, index=cluster_labels, columns=ct_labels)

for cluster in cluster_labels:
    our_top = top_genes[top_genes['cluster'] == cluster]
    our_gene_powers = dict(zip(our_top['gene'], our_top['power']))
    
    for cell_type, paper_df in paper_markers.items():
        # Sum of power scores for genes that appear in both our top genes
        # and the paper's marker list for this cell type
        shared = set(our_gene_powers.keys()) & set(paper_df['gene'].tolist())
        if shared:
            score = sum(our_gene_powers[g] for g in shared)
            overlap_matrix.loc[cluster, cell_type] = round(score, 3)

fig, ax = plt.subplots(figsize=(10, 8))
im = ax.imshow(overlap_matrix.values, cmap='Reds', aspect='auto')

ax.set_xticks(range(len(ct_labels)))
ax.set_xticklabels(ct_labels, rotation=45, ha='right', fontsize=11)
ax.set_yticks(range(len(cluster_labels)))
ax.set_yticklabels(cluster_labels, fontsize=11)

for i in range(len(cluster_labels)):
    for j in range(len(ct_labels)):
        val = overlap_matrix.values[i, j]
        ax.text(j, i, f'{val:.2f}',
                ha='center', va='center', fontsize=8,
                color='white' if val > overlap_matrix.values.max() * 0.6 else 'black')

plt.colorbar(im, ax=ax, label='Summed classification power of shared marker genes')
ax.set_title('ROC-based cluster correspondence\n(our clusters vs paper cell types)')
ax.set_xlabel('Paper cell types (Dataset S2)')
ax.set_ylabel('Our clusters')
fig.tight_layout()
fig.savefig("figures/roc_heatmap.png", dpi=150, bbox_inches='tight')
plt.close()
print("Saved: figures/roc_heatmap.png")

print("\nDone!")
