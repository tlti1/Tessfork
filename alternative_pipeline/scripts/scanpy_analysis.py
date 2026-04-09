import scanpy as sc
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl

# ── 1. Load featureCounts matrix ──────────────────────────────────────────────
print("Loading count matrix...")
df = pd.read_csv(
    "all_cells_counts.txt",
    sep="\t",
    comment="#",
    index_col=0
)

# Drop annotation columns, keep only count columns
df = df.drop(columns=["Chr", "Start", "End", "Strand", "Length", "gene_id"])

# Clean up column names
df.columns = [c.split("/")[-1].replace("_sorted.bam", "") for c in df.columns]
print(f"Matrix shape: {df.shape}")

# ── 2. Load metadata and filter to organoid cells only ───────────────────────
print("Loading metadata...")
meta = pd.read_csv("SraRunTable.csv", index_col="Run")

organoid_mask = meta["tissue"].isin([
    "Dissociated whole cerebral organoid",
    "Microdissected cortical-like ventricle from cerebral organoid"
])
organoid_srr = meta[organoid_mask].index.tolist()
organoid_srr = [s for s in organoid_srr if s in df.columns]
df = df[organoid_srr]
print(f"Organoid cells: {df.shape[1]}")

# ── 3. Create AnnData object ──────────────────────────────────────────────────
adata = sc.AnnData(df.T)
adata.obs = meta.loc[adata.obs_names, ["tissue", "stage", "source_name"]]
print(adata)

# ── 4. Filter cells: require ACTB and GAPDH expression ───────────────────────
print("Filtering cells by housekeeping genes...")
actb = adata[:, "ACTB"].X.flatten()
gapdh = adata[:, "GAPDH"].X.flatten()
adata = adata[(actb > 0) & (gapdh > 0)].copy()
print(f"Cells after housekeeping filter: {adata.n_obs}")

# ── 5. Basic QC filtering ─────────────────────────────────────────────────────
sc.pp.filter_cells(adata, min_genes=200)
sc.pp.filter_genes(adata, min_cells=3)
print(f"After QC filtering: {adata.shape}")

# ── 6. Normalisation and log transformation ───────────────────────────────────
sc.pp.normalize_total(adata, target_sum=1e4)
sc.pp.log1p(adata)

# ── 7. Load paper's exact PCA gene list from Dataset S2 ──────────────────────
print("Loading paper's PCA genes from Dataset S2...")
wb = openpyxl.load_workbook("pnas.1520760112.sd02.xlsx")
ws = wb['PCA_genes']
pca_genes = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
print(f"Paper's PCA genes: {len(pca_genes)}")

# Find which paper genes are in our dataset
pca_genes_present = [g for g in pca_genes if g in adata.var_names]
print(f"PCA genes found in our dataset: {len(pca_genes_present)}")

# Subset to paper's PCA genes
adata_pca = adata[:, pca_genes_present].copy()

# ── 8. Scale data ─────────────────────────────────────────────────────────────
sc.pp.scale(adata_pca, max_value=10)

# ── 9. PCA ────────────────────────────────────────────────────────────────────
print("Running PCA...")
sc.tl.pca(adata_pca, svd_solver="arpack")

sc.pl.pca_variance_ratio(adata_pca, log=True, show=False)
plt.savefig("figures/pca_variance_ratio.png")
plt.close()

# ── 10. Neighbourhood graph ───────────────────────────────────────────────────
sc.pp.neighbors(adata_pca, n_neighbors=10, n_pcs=30)

# ── 11. Leiden clustering ─────────────────────────────────────────────────────
sc.tl.leiden(adata_pca, resolution=1.5, flavor='igraph',
             n_iterations=2, directed=False)
print(f"Clusters found: {adata_pca.obs['leiden'].nunique()}")

# Copy clustering back to full adata
adata.obs['leiden'] = adata_pca.obs['leiden']

# ── 12. t-SNE with perplexity=5 (matching paper exactly) ─────────────────────
print("Running t-SNE...")
sc.tl.tsne(adata_pca, perplexity=5, random_state=0)
adata.obsm['X_tsne'] = adata_pca.obsm['X_tsne']

# ── 13. Plot t-SNE with colours for clusters and shapes for experiments ───────

# Create experiment label combining tissue + stage
adata.obs['experiment'] = (
    adata.obs['tissue'].str.replace('Dissociated whole cerebral organoid', 'Whole') 
    + ' ' + adata.obs['stage']
)

# Define markers for each experiment
experiment_markers = {
    'Whole 33 days': 'o',
    'Whole 35 days': 's',
    'Whole 37 days': '^',
    'Whole 41 days': 'D',
    'Whole 65 days': 'v',
    'Microdissected cortical-like ventricle from cerebral organoid 53 days': 'P',
    'Microdissected cortical-like ventricle from cerebral organoid 58 days': '*',
}

# Fix the microdissected labels
adata.obs['experiment'] = adata.obs['experiment'].str.replace(
    'Microdissected cortical-like ventricle from cerebral organoid', 'Microdissected'
)

experiment_markers = {
    'Whole 33 days': 'o',
    'Whole 35 days': 's',
    'Whole 37 days': '^',
    'Whole 41 days': 'D',
    'Whole 65 days': 'v',
    'Microdissected 53 days': 'P',
    'Microdissected 58 days': '*',
}

# Get cluster colours from scanpy palette
n_clusters = adata.obs['leiden'].nunique()
cluster_colors = sc.pl.palettes.default_20[:n_clusters]
color_map = {str(i): cluster_colors[i] for i in range(n_clusters)}

# Get t-SNE coordinates
tsne1 = adata.obsm['X_tsne'][:, 0]
tsne2 = adata.obsm['X_tsne'][:, 1]

fig, ax = plt.subplots(figsize=(10, 8))

# Plot each experiment with its own marker shape, coloured by cluster
for exp, marker in experiment_markers.items():
    mask = adata.obs['experiment'] == exp
    if mask.sum() == 0:
        continue
    colors = [color_map[c] for c in adata.obs['leiden'][mask]]
    ax.scatter(
        tsne1[mask], tsne2[mask],
        c=colors,
        marker=marker,
        s=40,
        label=exp,
        alpha=0.8,
        edgecolors='none'
    )

# Add cluster colour legend
from matplotlib.patches import Patch
cluster_legend = [Patch(color=cluster_colors[i], label=f'Cluster {i}')
                  for i in range(n_clusters)]
legend1 = ax.legend(handles=cluster_legend, title='Cluster',
                    loc='upper right', fontsize=7)
ax.add_artist(legend1)

# Add experiment shape legend
from matplotlib.lines import Line2D
shape_legend = [Line2D([0], [0], marker=m, color='grey', linestyle='None',
                        markersize=8, label=e)
                for e, m in experiment_markers.items()]
ax.legend(handles=shape_legend, title='Experiment',
          loc='lower right', fontsize=7)

ax.set_xlabel('tSNE1')
ax.set_ylabel('tSNE2')
ax.set_title('Organoid cells - t-SNE clusters (colours) by experiment (shapes)')
fig.savefig("figures/tsne_clusters_shapes.png", dpi=150, bbox_inches='tight')
plt.close()
print("Saved: figures/tsne_clusters_shapes.png")

# ── 14. Plot by tissue type (shapes for different experiments) ────────────────
fig, ax = plt.subplots(figsize=(8, 6))
sc.pl.tsne(adata, color="tissue", title="Organoid cells - tissue type",
           ax=ax, show=False)
fig.savefig("figures/tsne_organoid_tissue.png", dpi=150, bbox_inches='tight')
plt.close()
print("Saved: figures/tsne_organoid_tissue.png")

# ── 15. Plot key marker genes ─────────────────────────────────────────────────
marker_genes = [
    'SOX2', 'PAX6', 'VIM',
    'MKI67', 'ASPM',
    'EOMES',
    'TBR1', 'BCL11B',
    'TUBB3', 'MYT1L',
    'GAD1', 'FOXG1',
    'PECAM1', 'OTX2',
    'NFIA', 'NFIB'
]
markers_present = [g for g in marker_genes if g in adata.var_names]
print(f"Marker genes found: {markers_present}")

sc.pl.tsne(adata, color=markers_present, show=False, ncols=4)
plt.savefig("figures/tsne_marker_genes.png", dpi=150, bbox_inches='tight')
plt.close()
print("Saved: figures/tsne_marker_genes.png")

# ── 16. Dotplot of markers per cluster ───────────────────────────────────────
fig, ax = plt.subplots(figsize=(12, 6))
sc.pl.dotplot(adata, markers_present, groupby='leiden', ax=ax, show=False)
fig.savefig("figures/dotplot_markers.png", dpi=150, bbox_inches='tight')
plt.close()
print("Saved: figures/dotplot_markers.png")

print(f"\nDone! Found {adata.obs['leiden'].nunique()} clusters")

# ── 17. Match our clusters to paper's c1-c11 using marker genes ──────────────
print("\nChecking marker gene expression per cluster...")

# Key markers for each paper cluster
paper_markers = {
    'c1_c2_c3_cortical_NPC': ['FOXG1', 'NFIA', 'NFIB'],
    'c4_cortical_neuron': ['FOXG1', 'NEUROD6'],
    'c5_c6_noncortical_NPC': ['OTX2'],
    'c7_ventral_neuron': ['GAD1', 'DLX6'],
    'c8_c9_hem': ['WNT2B', 'RSPO3'],
    'c10_c11_mesenchyme': ['COL3A1', 'LUM'],
}

# Calculate mean expression per cluster for each marker
for cell_type, markers in paper_markers.items():
    present = [m for m in markers if m in adata.var_names]
    if not present:
        print(f"{cell_type}: none of {markers} found in dataset")
        continue
    print(f"\n{cell_type} markers {present}:")
    for cluster in sorted(adata.obs['leiden'].unique(), key=lambda x: int(x)):
        mask = adata.obs['leiden'] == cluster
        vals = adata[mask, present].X
        mean_expr = float(np.mean(vals))
        print(f"  Cluster {cluster}: mean expression = {mean_expr:.3f}")

# ── 18. Final t-SNE with cluster numbers directly on plot ─────────────────────
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.lines import Line2D
from scipy.spatial import ConvexHull

# Get t-SNE coordinates
tsne1 = adata.obsm['X_tsne'][:, 0]
tsne2 = adata.obsm['X_tsne'][:, 1]

# Get cluster colours from scanpy palette
n_clusters = adata.obs['leiden'].nunique()
cluster_colors = sc.pl.palettes.default_20[:n_clusters]
color_map = {str(i): cluster_colors[i] for i in range(n_clusters)}

# Create experiment label
adata.obs['experiment'] = (
    adata.obs['tissue'].astype(str)
    .str.replace('Dissociated whole cerebral organoid', 'Whole')
    .str.replace('Microdissected cortical-like ventricle from cerebral organoid', 'Microdissected')
    + ' ' + adata.obs['stage'].astype(str)
)

experiment_markers = {
    'Whole 33 days': 'o',
    'Whole 35 days': 's',
    'Whole 37 days': '^',
    'Whole 41 days': 'D',
    'Whole 65 days': 'v',
    'Microdissected 53 days': 'P',
    'Microdissected 58 days': '*',
}

fig, ax = plt.subplots(figsize=(12, 9))

# ── Draw smooth kernel density borders per cluster ────────────────────────────
from scipy.stats import gaussian_kde
import numpy as np

x_grid = np.linspace(tsne1.min() - 5, tsne1.max() + 5, 200)
y_grid = np.linspace(tsne2.min() - 5, tsne2.max() + 5, 200)
xx, yy = np.meshgrid(x_grid, y_grid)
grid_coords = np.vstack([xx.ravel(), yy.ravel()])

for i in range(n_clusters):
    cluster = str(i)
    mask = adata.obs['leiden'] == cluster
    if mask.sum() < 5:
        continue
    points = np.vstack([tsne1[mask], tsne2[mask]])
    try:
        kde = gaussian_kde(points, bw_method=0.4)
        zz = kde(grid_coords).reshape(xx.shape)
        # Find a contour level that captures most of the cluster
        threshold = zz.max() * 0.15
        ax.contourf(xx, yy, zz,
                    levels=[threshold, zz.max()],
                    colors=[cluster_colors[i]],
                    alpha=0.15,
                    zorder=0)
        ax.contour(xx, yy, zz,
                   levels=[threshold],
                   colors=[cluster_colors[i]],
                   alpha=0.5,
                   linewidths=1.5,
                   zorder=0)
    except Exception:
        pass
        
# Plot cells with shapes per experiment, colours per cluster
for exp, marker in experiment_markers.items():
    mask_exp = adata.obs['experiment'] == exp
    if mask_exp.sum() == 0:
        continue
    colors = [color_map[c] for c in adata.obs['leiden'][mask_exp]]
    ax.scatter(
        tsne1[mask_exp], tsne2[mask_exp],
        c=colors,
        marker=marker,
        s=40,
        alpha=0.8,
        edgecolors='none',
        zorder=1
    )

# ROC-based biological labels (from formal ROC test)
# C6 and C11 left as Unassigned due to low/ambiguous ROC signal
roc_labels = {
    'C1': 'Cycling cells',
    'C2': 'Ventral NPC',
    'C3': 'Cycling cells',
    'C4': 'Neurons',
    'C5': 'Hem/RSPO',
    'C6': 'Unassigned',
    'C7': 'Mesenchyme',
    'C8': 'Cycling progenitors',
    'C9': 'Dorsal NPC',
    'C10': 'Cortical neurons',
    'C11': 'Unassigned',
}

# Add cluster number labels directly on plot at centroid
# Add ROC-based biological labels directly on plot at centroid
for i in range(n_clusters):
    cluster = str(i)
    cluster_label = f'C{int(cluster)+1}'
    bio_label = roc_labels.get(cluster_label, 'Unassigned')
    mask = adata.obs['leiden'] == cluster
    x = tsne1[mask].mean()
    y = tsne2[mask].mean()
    # Two line label: cluster number + biological identity
    label_text = f'{cluster_label}\n{bio_label}'
    ax.annotate(
        label_text,
        (x, y),
        fontsize=8,
        fontweight='bold',
        ha='center',
        va='center',
        bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                  alpha=0.8, edgecolor='grey', linewidth=0.5),
        zorder=2
    )
# Experiment shape legend only
shape_legend = [
    Line2D([0], [0], marker=m, color='grey', linestyle='None',
           markersize=8, label=e)
    for e, m in experiment_markers.items()
]
ax.legend(handles=shape_legend, title='Experiment',
          loc='lower right', fontsize=8)

ax.set_xlabel('tSNE1')
ax.set_ylabel('tSNE2')
ax.set_title('Figure 3D replication - Organoid cell clusters\n(labels assigned by ROC test, unassigned = ambiguous signal)')
fig.savefig("figures/tsne_final.png", dpi=150, bbox_inches='tight')
plt.close()
print("Saved: figures/tsne_final.png")

# ── 19. Dotplot showing paper marker genes per cluster ────────────────────────
# Key markers from Dataset S2 representing each broad cell type
marker_genes_ordered = [
    'FOXG1', 'NFIA', 'NFIB',           # Cortical NPCs
    'NEUROD6', 'BCL11A', 'DCX',         # Cortical neurons
    'OTX2', 'FABP7', 'BCAT1',          # Non-cortical NPCs
    'GAD1', 'DLX6',                     # Ventral neurons
    'WNT2B', 'RSPO2', 'RSPO3', 'WLS',  # Hem cells
    'COL3A1', 'LUM', 'DCN', 'SPARC',   # Mesenchyme
]

markers_present = [g for g in marker_genes_ordered if g in adata.var_names]

# Rename leiden clusters to C1-C11 for display
adata.obs['cluster'] = adata.obs['leiden'].apply(lambda x: f'C{int(x)+1}')

sc.pl.dotplot(
    adata,
    markers_present,
    groupby='cluster',
    show=False,
    title='Marker gene expression per cluster'
)
plt.savefig("figures/dotplot_final.png", dpi=150, bbox_inches='tight')
plt.close()
print("Saved: figures/dotplot_final.png")

print(f"\nDone! {n_clusters} clusters identified.")
print("Cluster biological correspondence can be assessed from dotplot.")

# ── 20. Comparison plots: our data vs paper ───────────────────────────────────
import openpyxl
from scipy.stats import spearmanr

# Load paper marker genes with AUC scores from Dataset S2
wb = openpyxl.load_workbook("pnas.1520760112.sd02.xlsx")

paper_cell_types = {
    'Neurons': 'neurons_all',
    'Dorsal NPC': 'dorsal_cortex_progenitors',
    'Dorsal neuron': 'dorsal_cortex_neurons',
    'Ventral NPC': 'ventral_progenitors',
    'Hem/RSPO': 'RSPO_cells',
    'Mesenchyme': 'mesenchyme_cells',
}

# Get top 5 marker genes per cell type from paper
paper_top_genes = {}
for label, sheet in paper_cell_types.items():
    ws = wb[sheet]
    genes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[3]:
            genes.append((row[0], row[3]))  # gene, power
    # Sort by power, take top 5
    genes = sorted(genes, key=lambda x: x[1], reverse=True)[:5]
    paper_top_genes[label] = [g[0] for g in genes]

print("\nPaper top 5 genes per cell type:")
for ct, genes in paper_top_genes.items():
    print(f"  {ct}: {genes}")

# ── Dotplot comparing our clusters to paper marker genes ─────────────────────
# Flatten all paper marker genes into ordered list
all_paper_genes = []
gene_labels = []
for ct, genes in paper_top_genes.items():
    for g in genes:
        if g in adata.var_names and g not in all_paper_genes:
            all_paper_genes.append(g)
            gene_labels.append(ct)

print(f"\nTotal unique paper marker genes found in our data: {len(all_paper_genes)}")

# Plot dotplot with paper's marker genes grouped by cell type
sc.pl.dotplot(
    adata,
    all_paper_genes,
    groupby='cluster',
    show=False,
    title='Our clusters vs Paper marker genes (Dataset S2)',
    var_group_positions=[(0, 4), (5, 9), (10, 14), (15, 19), (20, 24), (25, 29)],
    var_group_labels=list(paper_cell_types.keys()),
)
plt.savefig("figures/dotplot_comparison.png", dpi=150, bbox_inches='tight')
plt.close()
print("Saved: figures/dotplot_comparison.png")

# ── Heatmap: correlation between our clusters and paper cell types ────────────
# For each cluster, compute mean expression of each paper cell type's markers
# Then correlate to get a similarity score

# Build mean expression matrix: clusters x paper cell types
cluster_ids = sorted(adata.obs['cluster'].unique(),
                     key=lambda x: int(x[1:]))

corr_matrix = pd.DataFrame(index=cluster_ids,
                            columns=list(paper_cell_types.keys()),
                            dtype=float)

for cluster in cluster_ids:
    mask = adata.obs['cluster'] == cluster
    for ct, genes in paper_top_genes.items():
        present = [g for g in genes if g in adata.var_names]
        if not present:
            corr_matrix.loc[cluster, ct] = 0
            continue
        mean_expr = float(adata[mask, present].X.mean())
        corr_matrix.loc[cluster, ct] = mean_expr

corr_matrix = corr_matrix.astype(float)

# Plot heatmap
fig, ax = plt.subplots(figsize=(10, 8))
im = ax.imshow(corr_matrix.values, cmap='Reds', aspect='auto')

ax.set_xticks(range(len(paper_cell_types)))
ax.set_xticklabels(list(paper_cell_types.keys()), rotation=45, ha='right')
ax.set_yticks(range(len(cluster_ids)))
ax.set_yticklabels(cluster_ids)

# Add value annotations
for i in range(len(cluster_ids)):
    for j in range(len(paper_cell_types)):
        val = corr_matrix.values[i, j]
        ax.text(j, i, f'{val:.2f}', ha='center', va='center',
                fontsize=8,
                color='white' if val > corr_matrix.values.max() * 0.6 else 'black')

plt.colorbar(im, ax=ax, label='Mean log-normalised expression')
ax.set_title('Similarity of our clusters to paper cell types\n(mean expression of paper marker genes)')
ax.set_xlabel('Paper cell types (Dataset S2)')
ax.set_ylabel('Our clusters')
fig.tight_layout()
fig.savefig("figures/heatmap_comparison.png", dpi=150, bbox_inches='tight')
plt.close()
print("Saved: figures/heatmap_comparison.png")




